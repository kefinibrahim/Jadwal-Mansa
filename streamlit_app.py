import streamlit as st
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO
import re

# Path folder xlsx untuk menyimpan file terjemahan
output_folder_path = "temp"

# Membuat database kode mata pelajaran
database_mata_pelajaran = {
    '2': 'Bahasa Inggris',
    '3': 'Fisika',
    '4': 'Sosiologi',
    '5': 'Bimbingan Konseling',
    '6': 'Matematika',
    '7': 'Bahasa Inggris',
    '8': 'Matematika',
    '9': 'Biologi',
    '10': 'Fisika',
    '11': 'Bahasa Inggris',
    '12': 'Kimia/Pkwu',
    '13': 'Kimia/Pkwu',
    '14': 'Bahasa Inggris',
    '15': 'Akidah Akhlak',
    '16': 'Informatika',
    '17': 'PPKN',
    '18': 'Bahasa Perancis',
    '19': 'PPKN',
    '20': 'Matematika',
    '21': 'Sejarah Indonesia',
    '22': 'Penjas',
    '23': 'Bahasa Indonesia',
    '24': 'Ekonomi/Pkwu',
    '25': 'Bahasa Indonesia',
    '26': 'Bimbingan Konseling',
    '27': 'Bahasa Arab',
    '28': 'Antropologi/Sosiologi',
    '29': 'Ekonomi/Pkwu',
    '30': 'Geografi',
    '31': 'Bahasa Arab',
    '32': 'Akidah Akhlak',
    '33': 'Bahasa Indonesia',
    '34': 'Kimia',
    '35': 'Quran Hadist',
    '36': 'Biologi',
    '37': 'Sejarah Indonesia',
    '38': 'Bahasa Indonesia',
    '39': 'Penjas',
    '40': 'Fikih/Ushul Fikih',
    '41': 'Seni Budaya',
    '42': 'Matematika',
    '43': 'SKI',
    '44': 'Bahasa Arab',
    '45': 'SKI',
    '46': 'Sejarah Indonesia',
    '47': 'Matematika',
    '48': 'Sejarah Indonesia/Riset',
    '49': 'Bimbingan Konseling',
    '50': 'Geografi/Riset',
    '51': 'Informatika',
    '52': 'Quran Hadist',
    '53': 'Bahasa Jawa',
    '54': 'Bimbingan Konseling',
    '55': 'Quran Hadist',
    '56': 'Matematika',
    '57': 'Bimbingan Konseling',
    '58': 'Penjas',
    '59': 'Fikih',
    '60': 'Tahfiz/Fikih',
    '61': 'Tahfiz',
    '62': 'Bahasa Arab Minat'
}

# Fungsi untuk menerjemahkan jadwal
def translate_jadwal(file_path):
    # Membaca jadwal dari file Excel menggunakan openpyxl
    wb = load_workbook(file_path)
    sheet = wb.active

    # Definisikan regular expression untuk mencocokkan kode guru
    kode_guru_regex = re.compile(r'\d{1,2}')

    # Menerjemahkan kode mata pelajaran
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row):
            if sheet.cell(row=1, column=cell.column).value == "Jam":
                # Kolom "Jam" tidak diterjemahkan
                continue
            cell_value = cell.value
            if cell_value:
                # Mencocokkan kode guru menggunakan regular expression
                kode_guru_match = kode_guru_regex.findall(str(cell_value))
                if kode_guru_match:
                    # Menggabungkan kode guru yang ditemukan menjadi satu string
                    mata_pelajaran = "/".join(database_mata_pelajaran.get(kode, '') for kode in kode_guru_match)
                    cell.value = mata_pelajaran

    # Mengatur format sel agar rapi
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width
        for cell in column_cells:
            cell.alignment = Alignment(wrapText=True)

    # Menyimpan jadwal terjemahan ke file Excel sementara
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# Judul halaman Streamlit
st.title("Jadwal mansa translator")
st.write("Konversi file pdf jadwal terlebih dahulu menjadi excel melalui https://www.pdf2go.com/pdf-to-excel, "
         "setelah itu baru masukkan ke dalam website ini! Jika suatu waktu kode guru berubah, maka website ini sudah tidak lagi valid -kefin ibrahim")

# Pengguna memilih file Excel untuk diterjemahkan
uploaded_file = st.file_uploader("Unggah File Excel", type=["xlsx"])

# Jika file Excel diunggah
if uploaded_file is not None:
    # Menjalankan terjemahan jadwal
    translated_file = translate_jadwal(uploaded_file)

    # Pengguna dapat mengunduh file terjemahan
    st.download_button("Unduh Hasil Terjemahan", data=translated_file, file_name=uploaded_file.name)
